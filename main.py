import os
import asyncio
import tkinter as tk
from tkinter.filedialog import askopenfilename
import traceback

from rich import print
from rich.console import Console
import openpyxl as opx
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
import aiohttp


async def check_uptime(path: str) -> None:
    red_fill = PatternFill(
        fill_type='solid',
        fgColor=Color('FFC7CE')
    )
    green_fill = PatternFill(
        fill_type='solid',
        fgColor=Color('C6EFCE')
    )

    console = Console()
    with console.status('Leyendo Excel...\n', spinner='point'):
        src_excel = opx.load_workbook(path)
    src_sheet = src_excel.active
    src_sheet.insert_cols(11)

    headers = src_sheet[7]
    notfound_excel = opx.Workbook()
    notfound_sheet = notfound_excel.active
    notfound_sheet.title = 'Contratos no encontrados'
    notfound_sheet.append(cell.value for cell in headers)

    session = aiohttp.ClientSession()
    for row in src_sheet.iter_rows(min_row=8):
        name = ' '.join(str(row[i].value) for i in range(5, 8, 1))
        url = row[9].value

        if not isinstance(url, str):
            break

        if not name.isupper():
            name = 'Clasificado'
        print(f'Contrato de [b]{name}[/]')

        with console.status(f'Checando enlace [cyan]{url}[/]...', spinner='point'):
            async with session.head(url) as response:
                status = response.status

        is_success = 200 <= status < 300
        color = '[green]' if is_success else '[red]'
        print(f'Código de estado: {color}{status}[/]\n')

        fill = green_fill if is_success else red_fill
        row[10].value = status
        row[9].fill = fill
        row[10].fill = fill

        if not is_success:
            notfound_sheet.append(cell.value for cell in row)

    await session.close()

    src_filename = 'estados.xlsx'
    with console.status(
            f'Guardando archivo con códigos de estado como [blue]{src_filename}[/]...',
            spinner='point'):
        src_excel.save(src_filename)

    notfound_filename = 'no_encontrados.xlsx'
    with console.status(
            f'Guardando archivo con contratos no encontrados como [magenta]{notfound_filename}[/]...',
            spinner='point'):
        notfound_excel.save(notfound_filename)

    with console.status(f'Abriendo [magenta]{notfound_filename}[/]...', spinner='point'):
        try:
            os.system(f"start excel.exe {notfound_filename}")
        except:
            print('Ocurrió un error al abrir Excel, pero el archivo se guardó correctamente.')


def main() -> None:
    console = Console()
    with console.status('Selecciona el archivo de Excel', spinner='point'):
        tk.Tk().withdraw()
        path = askopenfilename()
    
    if path == '':
        return

    print(f'Ruta del archivo: [yellow]{path}[/]')

    try:
        loop = asyncio.get_event_loop()
        loop.run_until_complete(check_uptime(path))
    except Exception:
        traceback.print_exc()
    finally:
        os.system('pause')


if __name__ == '__main__':
    main()
